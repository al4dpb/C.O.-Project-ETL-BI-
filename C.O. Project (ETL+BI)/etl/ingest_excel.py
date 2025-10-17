"""Excel ingestion to Bronze Parquet files - V2 with append-only partitions."""
import argparse
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Optional
from datetime import datetime

from etl.config import (
    EXCEL_PATH,
    BRONZE_DIR,
    COLUMN_ALIASES,
    DASHBOARD_SHEET_PATTERNS,
    LEASE_RATE_SHEET_PATTERNS,
    EXPENSE_SHEET_PATTERNS,
    OWN_USE_PATTERNS,
    VACANCY_PATTERNS,
    BRONZE_PARTITIONS,
    INGESTION_LOG_PATH,
)
from etl.utils import (
    logger,
    normalize_columns,
    pick_sheet,
    clean_dataframe,
    coerce_numeric,
    is_vacant,
    is_own_use,
    extract_building,
    compute_file_md5,
    detect_as_of_month,
    check_already_ingested,
    log_ingestion,
)


def ingest_dashboard_monthly(
    excel_path: str, as_of_month: str, file_md5: str
) -> pd.DataFrame:
    """
    Extract monthly dashboard KPIs from Excel.

    Returns:
        DataFrame with columns: as_of_month, rent_base, collected, uncollected,
                                leased_sqft, price_per_sf_yr, _file_md5, _ingested_at
    """
    logger.info("Ingesting Dashboard sheet...")

    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = pick_sheet(wb, DASHBOARD_SHEET_PATTERNS)

    if not sheet_name:
        raise ValueError(f"Could not find dashboard sheet in {excel_path}")

    # Extract from known structure (row 29+ contains the data table)
    df = _extract_dashboard_structured(excel_path, sheet_name, as_of_month)

    # Add audit columns
    df["as_of_month"] = as_of_month
    df["_file_md5"] = file_md5
    df["_ingested_at"] = datetime.now().isoformat()

    logger.info(f"Extracted {len(df)} monthly records")
    return df


def _extract_dashboard_structured(excel_path: str, sheet_name: str, as_of_month: str) -> pd.DataFrame:
    """
    Extract dashboard data from structured Excel layout (known positions).

    The Dashboard sheet has data starting at row 29 with months as columns.
    Row 29: Headers (Collection Target, January, February, ...)
    Row 30: Rent Base values
    Row 32: Collected values
    Row 33: Uncollected values
    Row 35: Leased SqFt values
    Row 36: Price / SqFt values

    Args:
        excel_path: Path to Excel file
        sheet_name: Sheet name
        as_of_month: Month in YYYY-MM format for year extraction

    Returns:
        DataFrame with monthly dashboard metrics
    """
    logger.info("Extracting from Dashboard sheet (rows 29+)...")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    # Extract month headers (row 29, starting from column B/index 1)
    header_row = 29
    months_raw = []
    for cell in ws[header_row]:
        if cell.value and str(cell.value).strip():
            val = str(cell.value).strip()
            # Skip "Collection Target" and similar labels
            if val.lower() not in ["collection target", "target"]:
                months_raw.append(val)

    logger.info(f"Found {len(months_raw)} months: {months_raw}")

    # Extract metric rows
    metrics = {
        "rent_base": 30,
        "collected": 32,
        "uncollected": 33,
        "leased_sqft": 35,
        "price_per_sf_yr": 36,
    }

    # Build data structure
    # Months start at column 9 (I) in the Excel
    data = []
    for month_idx, month_name in enumerate(months_raw):
        row_data = {"month": month_name}

        # Extract values for each metric
        for metric_name, row_num in metrics.items():
            # Column index: 9 is where months start (column I)
            col_num = 9 + month_idx
            cell_value = ws.cell(row=row_num, column=col_num).value

            if cell_value is not None:
                try:
                    row_data[metric_name] = float(cell_value)
                except (ValueError, TypeError):
                    row_data[metric_name] = None
            else:
                row_data[metric_name] = None

        data.append(row_data)

    df = pd.DataFrame(data)

    # Convert month names to dates (extract year from as_of_month)
    year = as_of_month.split('-')[0]  # Get year from as_of_month
    month_map = {
        "january": f"{year}-01-01",
        "february": f"{year}-02-01",
        "march": f"{year}-03-01",
        "april": f"{year}-04-01",
        "may": f"{year}-05-01",
        "june": f"{year}-06-01",
        "july": f"{year}-07-01",
        "august": f"{year}-08-01",
        "september": f"{year}-09-01",
        "october": f"{year}-10-01",
        "november": f"{year}-11-01",
        "december": f"{year}-12-01",
    }

    df["month"] = df["month"].str.lower().map(month_map)
    df["month"] = pd.to_datetime(df["month"], errors="coerce")

    # Drop rows with invalid months
    df = df.dropna(subset=["month"])

    logger.info(f"Extracted {len(df)} monthly records")
    return df


def ingest_expenses_monthly(
    excel_path: str, as_of_month: str, file_md5: str
) -> Optional[pd.DataFrame]:
    """
    Extract monthly expenses from Dashboard sheet (rows 42-70).

    Expense data is embedded in the Dashboard sheet with:
    - Row 40: Month headers starting at column I (index 8)
    - Rows 42-70: Expense line items
    - Column E (index 4): Item description
    - Columns I onward: Monthly actual values

    Returns:
        DataFrame with columns: as_of_month, month, item, actual,
                                expense_category, _file_md5, _ingested_at
    """
    logger.info("Ingesting Expenses from Dashboard sheet...")

    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet_name = pick_sheet(wb, DASHBOARD_SHEET_PATTERNS)

        if not sheet_name:
            logger.warning("Could not find Dashboard sheet, skipping expenses")
            return None

        ws = wb[sheet_name]

        # Extract month headers from row 40, starting at column I (index 9, 1-based)
        months_raw = []
        month_start_col = 9  # Column I in 1-based indexing
        for col_idx in range(month_start_col, month_start_col + 12):
            cell_value = ws.cell(row=40, column=col_idx).value
            if cell_value and str(cell_value).strip():
                months_raw.append(str(cell_value).strip())

        logger.info(f"Found {len(months_raw)} expense months: {months_raw}")

        # Map month names to dates (extract year from as_of_month)
        year = int(as_of_month.split('-')[0])
        month_map = {
            "january": f"{year}-01-01",
            "february": f"{year}-02-01",
            "march": f"{year}-03-01",
            "april": f"{year}-04-01",
            "may": f"{year}-05-01",
            "june": f"{year}-06-01",
            "july": f"{year}-07-01",
            "august": f"{year}-08-01",
            "september": f"{year}-09-01",
            "october": f"{year}-10-01",
            "november": f"{year}-11-01",
            "december": f"{year}-12-01",
        }

        # Define expense categories based on row groupings
        expense_items = {
            # Fixed operating expenses (rows 42-50)
            42: ("Electricity", "fixed"),
            43: ("Internet", "fixed"),
            44: ("Yardi Breeze", "fixed"),
            45: ("Water", "fixed"),
            46: ("Pest Control", "fixed"),
            47: ("Supplies", "fixed"),
            48: ("Nexus trash disposal", "fixed"),
            49: ("Insurance", "fixed"),
            50: ("Property management", "fixed"),
            # Variable expenses (rows 53-57)
            53: ("Repairs", "variable"),
            54: ("Real Estate Agent", "variable"),
            55: ("A/C Maintenance", "variable"),
            56: ("Bank Fees", "variable"),
            57: ("Light fixtures maintenance", "variable"),
            # Other expenses (rows 60-68)
            60: ("Accrued Property tax", "other"),
            61: ("Other", "other"),
            66: ("Management - RP", "other"),
            67: ("Other expenses", "other"),
            68: ("CAPEX", "other"),
        }

        # Extract expense data
        data = []
        for row_num, (item_name, category) in expense_items.items():
            # Get item from column E (index 5, 1-based)
            item_cell = ws.cell(row=row_num, column=5).value
            item = str(item_cell).strip() if item_cell else item_name

            # Extract monthly values
            for month_idx, month_name in enumerate(months_raw):
                col_num = month_start_col + month_idx
                actual_value = ws.cell(row=row_num, column=col_num).value

                # Only add if there's a value
                if actual_value is not None:
                    try:
                        actual = float(actual_value)
                        month_date = month_map.get(month_name.lower())

                        if month_date:
                            data.append({
                                "month": month_date,
                                "item": item,
                                "actual": actual,
                                "expense_category": category,
                            })
                    except (ValueError, TypeError):
                        pass  # Skip non-numeric values

        if not data:
            logger.warning("No expense data extracted")
            return None

        df = pd.DataFrame(data)
        df["month"] = pd.to_datetime(df["month"])

        # Add metadata
        df["as_of_month"] = as_of_month
        df["_file_md5"] = file_md5
        df["_ingested_at"] = datetime.now().isoformat()

        # Final column order
        df = df[[
            "as_of_month",
            "month",
            "item",
            "actual",
            "expense_category",
            "_file_md5",
            "_ingested_at",
        ]]

        logger.info(f"Extracted {len(df)} expense records")
        logger.info(f"  - Fixed: {len(df[df['expense_category']=='fixed'])}")
        logger.info(f"  - Variable: {len(df[df['expense_category']=='variable'])}")
        logger.info(f"  - Other: {len(df[df['expense_category']=='other'])}")

        return df

    except Exception as e:
        logger.error(f"Error extracting expenses: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def ingest_lease_rate_snapshot(
    excel_path: str, as_of_month: str, file_md5: str
) -> pd.DataFrame:
    """
    Extract lease/rate snapshot from Excel.

    The Lease Rate sheet has headers in row 4:
    Suite | Tenant | Lease End-Date | Offices | Meeting Rooms | Avg Rent per Unit | Sq Ft | Monthly Rent ($) | Annual Rent ($)

    Returns:
        DataFrame with columns: as_of_month, suite_id, building, tenant, sqft,
                                rent_monthly, rent_annual, rent_psf_yr, is_vacant,
                                is_own_use, _file_md5, _ingested_at
    """
    logger.info("Ingesting Lease Rate sheet...")

    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = pick_sheet(wb, LEASE_RATE_SHEET_PATTERNS)

    if not sheet_name:
        raise ValueError(f"Could not find lease/rate sheet in {excel_path}")

    # Read with header at row 4 (3 in 0-based indexing)
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=3)
    df = clean_dataframe(df)

    logger.info(f"Lease Rate raw columns: {list(df.columns)}")

    # Rename columns to standard names
    column_map = {
        "Suite": "suite_id",
        "Tenant": "tenant",
        "Sq Ft": "sqft",
        "Monthly Rent ($)": "rent_monthly",
        "Annual Rent ($)": "rent_annual",
    }

    # Apply rename for columns that exist
    rename_dict = {k: v for k, v in column_map.items() if k in df.columns}
    df = df.rename(columns=rename_dict)

    logger.info(f"Lease Rate after rename: {list(df.columns)}")

    # Ensure required columns exist
    if "suite_id" not in df.columns:
        df["suite_id"] = [f"SUITE_{i+1}" for i in range(len(df))]

    # Fill missing suite_ids (happens with empty rows)
    df["suite_id"] = df["suite_id"].fillna("").astype(str)
    df = df[df["suite_id"].str.strip() != ""]  # Remove rows with empty suite_id

    if "tenant" not in df.columns:
        df["tenant"] = "Unknown"

    # Extract building from suite_id
    df["building"] = df["suite_id"].apply(extract_building)

    # Numeric columns
    for col in ["sqft", "rent_monthly", "rent_annual"]:
        if col in df.columns:
            df[col] = coerce_numeric(df[col])
        else:
            df[col] = 0.0

    # Calculate rent per sqft per year if not present
    df["rent_psf_yr"] = 0.0
    if "rent_annual" in df.columns and "sqft" in df.columns:
        df["rent_psf_yr"] = df.apply(
            lambda row: row["rent_annual"] / row["sqft"] if row["sqft"] > 0 else 0.0,
            axis=1,
        )

    # Derive flags
    df["is_vacant"] = df["tenant"].apply(lambda x: is_vacant(x, VACANCY_PATTERNS))
    df["is_own_use"] = df.apply(
        lambda row: is_own_use(
            str(row.get("tenant", "")),
            str(row.get("suite_id", "")),
            OWN_USE_PATTERNS,
        ),
        axis=1,
    )

    # Add audit columns
    df["as_of_month"] = as_of_month
    df["_file_md5"] = file_md5
    df["_ingested_at"] = datetime.now().isoformat()

    # Select final columns
    final_cols = [
        "as_of_month",
        "suite_id",
        "building",
        "tenant",
        "sqft",
        "rent_monthly",
        "rent_annual",
        "rent_psf_yr",
        "is_vacant",
        "is_own_use",
        "_file_md5",
        "_ingested_at",
    ]
    df = df[final_cols]

    logger.info(f"Extracted {len(df)} lease records")
    logger.info(f"  - Vacant: {df['is_vacant'].sum()}")
    logger.info(f"  - Own-use: {df['is_own_use'].sum()}")

    return df


def save_bronze_partition(
    df: pd.DataFrame, table_name: str, as_of_month: str
) -> None:
    """
    Save DataFrame to Bronze Parquet with as_of_month partition.

    Args:
        df: DataFrame to save
        table_name: Table name (key in BRONZE_PARTITIONS)
        as_of_month: Month in YYYY-MM format
    """
    partition_base = BRONZE_PARTITIONS[table_name]
    partition_path = partition_base / f"as_of_month={as_of_month}"
    partition_path.mkdir(parents=True, exist_ok=True)

    # Use the full bronze table name for the file (not the short table_name)
    bronze_table_name = partition_base.name  # e.g., "raw_dashboard_monthly"
    output_file = partition_path / f"{bronze_table_name}_{as_of_month}.parquet"

    df.to_parquet(output_file, index=False, engine="pyarrow")
    logger.info(f"✓ Saved {len(df)} rows to {output_file}")


def main(file_path: Optional[str] = None):
    """Run all ingestion tasks with append-only partitioning."""
    logger.info("=" * 60)
    logger.info("Starting Excel → Bronze ingestion (V2)")
    logger.info("=" * 60)

    started_at = datetime.now()

    # Determine file path
    if file_path is None:
        file_path = EXCEL_PATH
        logger.info(f"Using default Excel path: {file_path}")
    else:
        logger.info(f"Using provided Excel path: {file_path}")

    # Compute file MD5
    file_md5 = compute_file_md5(file_path)
    logger.info(f"File MD5: {file_md5}")

    # Detect as_of_month
    wb = openpyxl.load_workbook(file_path, data_only=True)
    dashboard_sheet = pick_sheet(wb, DASHBOARD_SHEET_PATTERNS)
    as_of_month = detect_as_of_month(file_path, dashboard_sheet)
    logger.info(f"Detected as_of_month: {as_of_month}")

    # Check if already ingested
    already_ingested = check_already_ingested(
        file_md5, as_of_month, BRONZE_PARTITIONS["dashboard"]
    )
    if already_ingested:
        logger.info("File already ingested, skipping")
        log_ingestion(INGESTION_LOG_PATH, file_md5, as_of_month, 0, started_at, "skipped")
        return

    total_rows = 0

    # Ingest dashboard monthly data
    try:
        df_dashboard = ingest_dashboard_monthly(file_path, as_of_month, file_md5)
        save_bronze_partition(df_dashboard, "dashboard", as_of_month)
        total_rows += len(df_dashboard)
    except Exception as e:
        logger.error(f"Failed to ingest dashboard: {e}")
        log_ingestion(INGESTION_LOG_PATH, file_md5, as_of_month, 0, started_at, "failed")
        raise

    # Ingest expenses monthly (optional)
    try:
        df_expenses = ingest_expenses_monthly(file_path, as_of_month, file_md5)
        if df_expenses is not None and len(df_expenses) > 0:
            save_bronze_partition(df_expenses, "expenses", as_of_month)
            total_rows += len(df_expenses)
    except Exception as e:
        logger.warning(f"Failed to ingest expenses: {e}")

    # Ingest lease/rate snapshot
    try:
        df_lease = ingest_lease_rate_snapshot(file_path, as_of_month, file_md5)
        save_bronze_partition(df_lease, "lease_rate", as_of_month)
        total_rows += len(df_lease)
    except Exception as e:
        logger.error(f"Failed to ingest lease rate: {e}")
        log_ingestion(INGESTION_LOG_PATH, file_md5, as_of_month, total_rows, started_at, "failed")
        raise

    # Log successful ingestion
    log_ingestion(INGESTION_LOG_PATH, file_md5, as_of_month, total_rows, started_at, "success")

    logger.info("=" * 60)
    logger.info(f"✓ Bronze ingestion completed successfully ({total_rows} total rows)")
    logger.info("=" * 60)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ingest CFO Excel to Bronze")
    parser.add_argument(
        "--file", type=str, help="Path to Excel file (default: from .env)"
    )
    args = parser.parse_args()

    main(file_path=args.file)
