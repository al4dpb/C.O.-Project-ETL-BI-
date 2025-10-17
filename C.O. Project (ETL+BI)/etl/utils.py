"""Utility functions for ETL processing."""
import logging
import re
import hashlib
from pathlib import Path
from typing import Any, List, Optional
from datetime import datetime
import pandas as pd
import openpyxl

# Setup logger
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def normalize_columns(df: pd.DataFrame, alias_map: dict) -> pd.DataFrame:
    """
    Normalize column names using alias map.

    Args:
        df: DataFrame with potentially variant column names
        alias_map: Dictionary mapping variants to standard names

    Returns:
        DataFrame with normalized column names
    """
    df = df.copy()

    # Normalize to lowercase and strip whitespace
    df.columns = df.columns.str.lower().str.strip()

    # Apply aliases
    df.columns = df.columns.map(lambda x: alias_map.get(x, x))

    logger.info(f"Normalized columns: {list(df.columns)}")
    return df


def pick_sheet(workbook, patterns: List[str]) -> Optional[str]:
    """
    Find sheet name matching any of the patterns (case-insensitive).

    Args:
        workbook: openpyxl Workbook object
        patterns: List of string patterns to match

    Returns:
        Sheet name or None if not found
    """
    sheet_names = [s.lower() for s in workbook.sheetnames]

    for pattern in patterns:
        pattern_lower = pattern.lower()
        for i, sheet_name in enumerate(sheet_names):
            if pattern_lower in sheet_name:
                actual_name = workbook.sheetnames[i]
                logger.info(f"Found sheet '{actual_name}' matching pattern '{pattern}'")
                return actual_name

    logger.warning(f"No sheet found matching patterns: {patterns}")
    return None


def coerce_numeric(series: pd.Series) -> pd.Series:
    """
    Coerce series to numeric, handling errors gracefully.

    Args:
        series: Pandas Series

    Returns:
        Numeric series with NaN for invalid values
    """
    return pd.to_numeric(series, errors='coerce')


def is_vacant(tenant_name: str, vacancy_patterns: List[str]) -> bool:
    """
    Check if tenant name indicates vacancy.

    Args:
        tenant_name: Tenant name string
        vacancy_patterns: List of patterns indicating vacancy

    Returns:
        True if vacant, False otherwise
    """
    if pd.isna(tenant_name) or not tenant_name:
        return True

    tenant_lower = str(tenant_name).lower().strip()
    return any(pattern in tenant_lower for pattern in vacancy_patterns)


def is_own_use(tenant_name: str, suite_id: str, own_use_patterns: List[str]) -> bool:
    """
    Check if suite is used by owner (non-revenue).

    Args:
        tenant_name: Tenant name string
        suite_id: Suite identifier
        own_use_patterns: List of patterns indicating owner use

    Returns:
        True if own-use, False otherwise
    """
    combined = f"{tenant_name} {suite_id}".lower().strip()
    return any(pattern in combined for pattern in own_use_patterns)


def extract_building(suite_id: str) -> Optional[str]:
    """
    Extract building (A or B) from suite ID.

    Args:
        suite_id: Suite identifier (e.g., "101A", "205B")

    Returns:
        'A' or 'B' or None
    """
    if pd.isna(suite_id):
        return None

    suite_str = str(suite_id).upper()
    match = re.search(r'[AB]', suite_str)
    return match.group() if match else None


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean DataFrame by removing empty rows/columns.

    Args:
        df: Input DataFrame

    Returns:
        Cleaned DataFrame
    """
    df = df.copy()

    # Drop completely empty rows and columns
    df = df.dropna(how='all')
    df = df.dropna(axis=1, how='all')

    # Strip whitespace from string columns
    str_cols = df.select_dtypes(include=['object']).columns
    df[str_cols] = df[str_cols].apply(lambda x: x.str.strip() if hasattr(x, 'str') else x)

    return df


def compute_file_md5(file_path: str) -> str:
    """
    Compute MD5 hash of file for deduplication.

    Args:
        file_path: Path to file

    Returns:
        MD5 hash as hex string
    """
    md5_hash = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            md5_hash.update(chunk)
    return md5_hash.hexdigest()


def detect_as_of_month(excel_path: str, dashboard_sheet_name: Optional[str] = None) -> str:
    """
    Detect as_of_month (YYYY-MM) from Excel file.

    Priority:
    1. Visible Month column in Dashboard sheet (first non-null value)
    2. "Period" cell in any sheet
    3. Filename pattern YYYY[-_]MM

    Args:
        excel_path: Path to Excel file
        dashboard_sheet_name: Name of dashboard sheet (optional)

    Returns:
        as_of_month in YYYY-MM format

    Raises:
        ValueError: If as_of_month cannot be detected
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Strategy 1: Dashboard sheet month column
    if dashboard_sheet_name and dashboard_sheet_name in wb.sheetnames:
        try:
            df = pd.read_excel(excel_path, sheet_name=dashboard_sheet_name, header=28)
            if 'month' in [str(c).lower() for c in df.columns]:
                month_col = [c for c in df.columns if str(c).lower() == 'month'][0]
                first_month = df[month_col].dropna().iloc[0]
                if pd.notna(first_month):
                    parsed = pd.to_datetime(first_month, errors='coerce')
                    if pd.notna(parsed):
                        as_of = parsed.strftime('%Y-%m')
                        logger.info(f"Detected as_of_month from Dashboard month column: {as_of}")
                        return as_of
        except Exception as e:
            logger.debug(f"Could not extract month from Dashboard sheet: {e}")

    # Strategy 2: Search for "Period" cell
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=50):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'period' in cell.value.lower():
                    # Check adjacent cells for date
                    for offset in [1, 2]:
                        try:
                            adjacent = ws.cell(row=cell.row, column=cell.column + offset).value
                            if adjacent:
                                parsed = pd.to_datetime(adjacent, errors='coerce')
                                if pd.notna(parsed):
                                    as_of = parsed.strftime('%Y-%m')
                                    logger.info(f"Detected as_of_month from Period cell: {as_of}")
                                    return as_of
                        except:
                            pass

    # Strategy 3: Filename pattern
    filename = Path(excel_path).stem
    match = re.search(r'(20\d{2})[-_](\d{2})', filename)
    if match:
        year, month = match.groups()
        as_of = f"{year}-{month}"
        logger.info(f"Detected as_of_month from filename: {as_of}")
        return as_of

    # Fallback: current month
    as_of = datetime.now().strftime('%Y-%m')
    logger.warning(f"Could not detect as_of_month, using current month: {as_of}")
    return as_of


def check_already_ingested(file_md5: str, as_of_month: str, partition_path: Path) -> bool:
    """
    Check if file with same MD5 and as_of_month was already ingested.

    Args:
        file_md5: MD5 hash of file
        as_of_month: Month in YYYY-MM format
        partition_path: Base partition path

    Returns:
        True if already ingested, False otherwise
    """
    month_partition = partition_path / f"as_of_month={as_of_month}"
    if not month_partition.exists():
        return False

    # Check parquet files for matching MD5 in metadata (if stored) or presence
    # For simplicity, check if partition exists with any data
    parquet_files = list(month_partition.glob("*.parquet"))
    if parquet_files:
        # Read one file to check for MD5 column
        try:
            sample = pd.read_parquet(parquet_files[0])
            if '_file_md5' in sample.columns:
                existing_md5s = sample['_file_md5'].unique()
                if file_md5 in existing_md5s:
                    logger.info(f"File MD5 {file_md5} already ingested for {as_of_month}")
                    return True
        except Exception as e:
            logger.debug(f"Could not check MD5 in parquet: {e}")

    return False


def log_ingestion(
    log_path: Path,
    file_md5: str,
    as_of_month: str,
    rows_ingested: int,
    started_at: datetime,
    status: str = "success"
):
    """
    Append ingestion record to log file.

    Args:
        log_path: Path to ingestion log CSV
        file_md5: MD5 hash of ingested file
        as_of_month: Month in YYYY-MM format
        rows_ingested: Number of rows ingested
        started_at: Start timestamp
        status: Status (success/failed/skipped)
    """
    finished_at = datetime.now()
    log_entry = pd.DataFrame([{
        '_file_md5': file_md5,
        'as_of_month': as_of_month,
        'rows_ingested': rows_ingested,
        'started_at': started_at.isoformat(),
        'finished_at': finished_at.isoformat(),
        'status': status,
    }])

    if log_path.exists():
        log_entry.to_csv(log_path, mode='a', header=False, index=False)
    else:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_entry.to_csv(log_path, mode='w', header=True, index=False)

    logger.info(f"Logged ingestion: {status} - {rows_ingested} rows for {as_of_month}")
